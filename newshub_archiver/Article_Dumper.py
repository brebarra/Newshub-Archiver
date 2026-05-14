import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from pathlib import Path

# Function fetching HTML content of website. Return error if failed
def fetch_html(url):
    response = requests.get(url)
    if response.status_code == 200:
        return response.text
    else:
        print("HTML fetch failed")
        return None

# Function reading HTML & extracting title, author
def extract_html(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
  
    if html_content:
       titles = soup.find_all(class_='c-ArticleHeading-title')
       authors = soup.find_all(class_='c-ArticleHeading-authorName')
       promo_tag = soup.find(class_='c-ArticleHeading-promoTag')
       author_names = []

       if authors:
           for author in authors:
               author_text = author.get_text(strip = True)
               author_names.append(author_text)
           return titles, [author_names]
        
       elif promo_tag and 'Breaking' in promo_tag.text:
            body_text = soup.find(class_='c-ArticleBody')
            if body_text:
                top_line = body_text.find(string=lambda text: 'By' in text)
                if top_line:
                    author_text = top_line.split('By', 1)[-1].strip()
                    return titles, [author_text]

       # No authors found in article
       return titles, []

    else:
        return [], []


# Function saving url, title and author to spreadsheet
def save_to_spreadsheet(filepath, url, titles, authors):
    
    # Create a new workbook if one doesn't exist
    try:
        wb = load_workbook(filepath)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1, value="URL")
    ws.cell(row=1, column=2, value="Title")
    ws.cell(row=1, column=3, value="Author")
    
    next_row = ws.max_row + 1
    
    # print("Authors - sts: ",authors) # Debug print
    
    for title, author_list in zip(titles, authors):
        ws.cell(row=next_row, column=1, value=url)
        ws.cell(row=next_row, column=2, value=title.text.strip() if title else "")
        
        # print("Author list: ",author_list) # Debug print
        
        if author_list: 
            if isinstance(author_list, list):
                ws.cell(row=next_row, column=3, value=', '.join(author_list))

            else:
                ws.cell(row=next_row, column=3, value=author_list)
                
        next_row += 1

    wb.save(filepath)


def main():
    BASE_DIR = Path(__file__).parent

    input_file = BASE_DIR / "extracted_links.txt"
    output_file = BASE_DIR / "website_data.xlsx"
    
    with open(input_file, 'r') as file:
        for line in file:
            url = line.strip()
            html_content = fetch_html(url)
            if html_content:
                titles, authors = extract_html(html_content)
                if authors:
                    save_to_spreadsheet(output_file, url, titles, authors)
                    print(f"Data saved for: {url}")
                else:
                    print(f"No author found for {url}")

main()